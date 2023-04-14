

from openpyxl import load_workbook

#处理Excel文件类
class HandleExcel:

    #file_path文件路径，sheet_name 子表sheet名称
    def __init__(self,file_path,sheet_name):

        self.wb = load_workbook(file_path)   #加载读取文件
        self.sh = self.wb[sheet_name]       #获取指定的工作表
        self.file_path = file_path

    #获取工作表第一行列名，以列表形式返回
    def __read_titles(self):
        titles = []
        for item in list(self.sh.rows)[0]:  # 遍历第1行当中每一列
            titles.append(item.value)
        return titles   #title列表存储列名

    #读取工作表所有除第一行外所有数据，并与列名打包成字典，存储在all_datas列表返回
    def read_all_datas(self):
        all_datas = []
        titles = self.__read_titles()
        for item in list(self.sh.rows)[1:]:  # 遍历数据行
            values = []
            for val in item:  # 获取每一行的值
                values.append(val.value)
            res = dict(zip(titles, values))  # title和每一行数据，打包成字典
            all_datas.append(res)
        return all_datas
    #写数据到单元格，行、列、值
    def write_data(self,row,column,value):
        #写入数据
        self.sh.cell(row,column,value)
        #保存数据
        self.wb.save(self.file_path)

    def close_file(self):
        self.wb.close()


if __name__ == '__main__':
    import os
    file_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "TestDatas\\AIops测试用例.xlsx")
    print(file_path)
    exc = HandleExcel(file_path,"监控管理")
    cases = exc.read_all_datas()
    print(cases)
    exc.close_file()
    for case in cases:
       print(case)

